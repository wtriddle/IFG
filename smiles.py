import ifg
import re
from openpyxl import Workbook
# import collect
# import evalute

# Openpyxl objects
fgWorkbook = Workbook()
fgContaimentSheet = fgWorkbook.active
fgAllSheet = fgWorkbook.create_sheet()

# Set intial column names
fgContaimentSheet.cell(row=1, column=1).value = "Refcode"
fgContaimentSheet.cell(row=1, column=2).value = "SMILES"
fgAllSheet.cell(row=1, column=1).value = "Refcode"
fgAllSheet.cell(row=1, column=2).value = "SMILES"

# List which holds functional group names
functionalGroups = []

# Grab all names from FGlist.txt
for line in open('FGlist.txt', 'r'):
    lineInfo = re.compile(r'\S+').findall(line)
    groupName = lineInfo[1]
    if groupName not in functionalGroups:
        functionalGroups.append(groupName)

# Append groups which are not found by FGlist.txt, but by direct evalution instead
functionalGroups.append("Alcohol")
functionalGroups.append("Acetal")
functionalGroups.append("Hemiketal")
functionalGroups.append("Hemiacetal")
functionalGroups.append("PrimaryAmine")

# Sort them alphabetically
functionalGroups.sort()
functionalGroups.insert(2,"totalAlcohols") # Special varibale that requires tracking
holder = functionalGroups.copy()

# Add cyclic/aromatic distinctions to list
indexCounter = insertions = 0
for name in holder:
    indexCounter += 1
    functionalGroups.insert(indexCounter+insertions, "Cyclic" + name)
    if len(re.compile(r'Amine').findall(name)) != 0 or name == "Alcohol":
        functionalGroups.insert(indexCounter+insertions, "Aromatic" + name)
        insertions += 1 # offset counter
    insertions += 1

# Insert functional group names into the top row of excel sheet
for column in range(3, len(functionalGroups)+3):
    fgContaimentSheet.cell(row=1,column=column).value = functionalGroups[column-3]
    fgAllSheet.cell(row=1, column=column).value = functionalGroups[column-3]

# Insert more special varibales into top of excel sheet
fgContaimentSheet.cell(row=1, column=len(functionalGroups)+3).value = "aromaticRingCount"
fgContaimentSheet.cell(row=1, column=len(functionalGroups)+4).value = "nonAromaticRingCount"
fgContaimentSheet.cell(row=1, column=len(functionalGroups)+5).value = "RingCount"
fgContaimentSheet.cell(row=1, column=len(functionalGroups)+6).value = "AminoAcid"

# Initializations
rowCounter = 1
maxContaimentColumn = fgContaimentSheet.max_column
maxFGallColumn = fgAllSheet.max_column

# Loop thorugh each structure SMILEScode
for line in open('smiles.txt', 'r'):
    rowCounter += 1

    # Insert 0's in all possible cells
    for column in range(2, maxContaimentColumn+1):
        fgContaimentSheet.cell(row=rowCounter, column=column).value = 0
    for column in range(2, maxFGallColumn+1):
        fgAllSheet.cell(row=rowCounter, column=column).value = 0

    # Retrieve SMILEScode/Refcode information, place into excel sheet in current row
    lineInfo = re.compile(r'\S+').findall(line)
    smiles = lineInfo[1]
    RefCode = lineInfo[2]
    fgContaimentSheet.cell(row=rowCounter, column=1).value = RefCode
    fgContaimentSheet.cell(row=rowCounter, column=2).value = smiles
    fgAllSheet.cell(row=rowCounter, column=1).value = RefCode
    fgAllSheet.cell(row=rowCounter, column=2).value = smiles

    # Evaluate the SMILEScode for its functional groups
    print("EVALUATING ", lineInfo[2], " ", smiles)
    functionalGroupData = ifg.ifg(smiles)
    print(functionalGroupData)

    # Place the resultant functional groups from dictionary object into the proper column/row name
    for group in functionalGroupData[1].items(): # Loop through the detemrined functional groups from SMILEScode
         for column in range(2,maxContaimentColumn+1): # Loop through every column...
            if fgContaimentSheet.cell(row=1, column=column).value == group[0]: # ...To find which column to place the data into
                fgContaimentSheet.cell(row=rowCounter, column=column).value = int(group[1]) # Insert the value into that row
                # Special cases for AminoAcids
                if group[0] == "AminoAcid" and int(group[1]) == 1:
                    fgContaimentSheet.cell(row=rowCounter, column=column).value = "Yes"
                elif group[0] == "AminoAcid" and int(group[1]) == 0:
                    fgContaimentSheet.cell(row=rowCounter, column=column).value = "No"
                break
    # Repeat loop revised set of data, minus special cases
    for group in functionalGroupData[3].items():
         for column in range(2,maxFGallColumn+1):
            if fgAllSheet.cell(row=1, column=column).value == group[0]:
                fgAllSheet.cell(row=rowCounter, column=column).value = int(group[1])
                break
    del(functionalGroupData)
fgWorkbook.save("functionalGroupData.xlsx") # Save the data sheet

# data = ifg.ifg("Nc1ccc(cc1)C(=O)OCCCOC(=O)c1ccc(N)cc1")
# for thing in data:
#     print("\n")
#     print(thing)
# print("Nc1ccc(cc1)C(=O)OCCCOC(=O)c1ccc(N)cc1")
