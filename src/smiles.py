from Molecule import Molecule
from ifg import ifg
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def createFgDataDict(functionalGroups):
    fgDataDict = {}
    for group in functionalGroups:
        if group.NAME in fgDataDict.keys():
            fgDataDict[group.NAME] += 1
        else:
            fgDataDict.update({group.NAME: 1})
    return fgDataDict


# Setup sheets
fgWorkbook = Workbook()
fgPreciseSheet = fgWorkbook.create_sheet('Precise Functional Groups')
fgAllSheet = fgWorkbook.create_sheet('All Functional Groups')
fgPreciseSheet.title = 'Precise Functional Groups'
fgAllSheet.title = 'All Functional Groups'

fgAllSheet.cell(row=1, column=1).value = "Refcodes"
fgAllSheet.cell(row=1, column=2).value = "Smiles"

fgPreciseSheet.cell(row=1, column=1).value = "Refcodes"
fgPreciseSheet.cell(row=1, column=2).value = "Smiles"


# Retrieve functional group names
functionalGroups = []
for line in open('./resources/FGlist.txt', 'r'):
    lineInfo = re.compile(r'\S+').findall(line)
    if lineInfo[1] not in functionalGroups:
        functionalGroups.append(lineInfo[1])

functionalGroups.append("Alcohol")
functionalGroups.append("PrimaryAmine")

# Sort and add cyclic/aromaic nomenclatures
functionalGroups.sort()
cyclicGroups = []
aromaticGroups = []
for group in functionalGroups:
    cyclicGroups.append("Cyclic" + group)
    aromaticGroups.append("Aromatic" + group)

# Input into names first row
col = 0
index = -1
while index != len(functionalGroups) - 1:
    print("adding ", functionalGroups[index])
    index += 1
    col += 3
    fgAllSheet.cell(row=1, column=col).value = functionalGroups[index]
    fgAllSheet.cell(row=1, column=col+1).value = cyclicGroups[index]
    fgAllSheet.cell(row=1, column=col+2).value = aromaticGroups[index]
    fgPreciseSheet.cell(row=1, column=col).value = functionalGroups[index]
    fgPreciseSheet.cell(row=1, column=col+1).value = cyclicGroups[index]
    fgPreciseSheet.cell(row=1, column=col+2).value = aromaticGroups[index]

# Input ring info into row 1
col += 2
ringInfo = ['aromaticRings', 'nonAromaticRings', 'ringCount']
index = -1
while index != len(ringInfo) - 1:
    index += 1
    col += 1
    fgAllSheet.cell(row=1, column=col).value = ringInfo[index]
    fgPreciseSheet.cell(row=1, column=col).value = ringInfo[index]

fgAllSheet.cell(row=1, column=col+1).value = "totalAlcohols"
fgPreciseSheet.cell(row=1, column=col+1).value = "totalAlcohols"
fgAllSheet.cell(row=1, column=col+2).value = "AminoAcid"
fgPreciseSheet.cell(row=1, column=col+2).value = "AminoAcid"

# Loop over smiles codes and find their FG/Ring data
row = 1
maxCol = fgAllSheet.max_column
for line in open('./resources/smiles.txt', 'r'):
    row += 1
    lineInfo = re.compile(r'\S+').findall(line)
    smiles = lineInfo[1]
    refcode = lineInfo[2]
    fgAllSheet.cell(row=row, column=1).value = refcode
    fgAllSheet.cell(row=row, column=2).value = smiles
    fgPreciseSheet.cell(row=row, column=1).value = refcode
    fgPreciseSheet.cell(row=row, column=2).value = smiles

    # Set current row to all 0's
    for col in range(3, maxCol):
        fgAllSheet.cell(row=row, column=col).value = 0
        fgPreciseSheet.cell(row=row, column=col).value = 0

    print(refcode)
    print(smiles)

    # Get the functional groups and ring data from ifg algorithm
    fgs = ifg(smiles, refcode)
    fgAllDict = createFgDataDict(fgs.functionalGroups)
    fgPreciseDict = createFgDataDict(fgs.preciseFunctionalGroups)
    totalAlcohols = len(fgs.ALCOHOLICINDICES)

    # Loop over data dictionaries and apply values to corellating columns
    for fg in fgAllDict.items():
        for col in range(3, maxCol):
            if fgAllSheet.cell(row=1, column=col).value == fg[0]:
                fgAllSheet.cell(row=row, column=col).value = fg[1]
                break

    for fg in fgPreciseDict.items():
        for col in range(3, maxCol):
            if fgPreciseSheet.cell(row=1, column=col).value == fg[0]:
                fgPreciseSheet.cell(row=row, column=col).value = fg[1]
                break

    for ring in fgs.RINGDICT.items():
        for col in range(3, maxCol):
            if fgPreciseSheet.cell(row=1, column=col).value == ring[0]:
                fgPreciseSheet.cell(row=row, column=col).value = ring[1]
                break

    for ring in fgs.RINGDICT.items():
        for col in range(3, maxCol):
            if fgAllSheet.cell(row=1, column=col).value == ring[0]:
                fgAllSheet.cell(row=row, column=col).value = ring[1]
                break

    fgAllSheet.cell(row=row, column=maxCol-1).value = totalAlcohols
    fgPreciseSheet.cell(row=row, column=maxCol-1).value = totalAlcohols

    if fgs.AMINOACID:
        fgAllSheet.cell(row=row, column=maxCol).value = "Yes"
        fgPreciseSheet.cell(row=row, column=maxCol).value = "Yes"
    else:
        fgAllSheet.cell(row=row, column=maxCol).value = "No"
        fgPreciseSheet.cell(row=row, column=maxCol).value = "No"

    print(fgAllDict)
    print(fgPreciseDict)
    print(fgs.RINGDICT)

# Get column indices with all 0's
index = 2
removableCols = []
for value in fgAllSheet.iter_cols(min_row=1, min_col=3, max_row=fgAllSheet.max_row, max_col=fgAllSheet.max_column, values_only=True):

    index += 1

    for val in value[1:len(value)]:
        if val != 0:
            break
    else:
        removableCols.append(index)

# Remove columns with all 0's
colShifter = 0
for col in removableCols:
    fgToRemove = fgAllSheet.cell(row=1, column=col-colShifter).value
    fgAllSheet.delete_cols(col-colShifter)
    colShifter += 1

del(removableCols)

# Get column indices with all 0's
index = 2
removableCols = []
for value in fgPreciseSheet.iter_cols(min_row=1, min_col=3, max_row=fgPreciseSheet.max_row, max_col=fgPreciseSheet.max_column, values_only=True):

    index += 1

    for val in value[1:len(value)]:
        if val != 0:
            break
    else:
        removableCols.append(index)

# Remove columns with all 0's
colShifter = 0
for col in removableCols:
    fgToRemove = fgPreciseSheet.cell(row=1, column=col-colShifter).value
    fgPreciseSheet.delete_cols(col-colShifter)
    colShifter += 1

# Freeze Refcodes, SMILES codes, and Functional groups
freeze1 = fgAllSheet['C2']
fgAllSheet.freeze_panes = freeze1
freeze2 = fgPreciseSheet['C2']
fgPreciseSheet.freeze_panes = freeze2

# Set up column widths for readability
for col in range(3, fgAllSheet.max_column+1):
    colWidth = len(fgAllSheet.cell(row=1, column=col).value)
    letterCol = get_column_letter(col)
    fgAllSheet.column_dimensions[letterCol].width = colWidth + 5

for col in range(3, fgPreciseSheet.max_column+1):
    colWidth = len(fgPreciseSheet.cell(row=1, column=col).value)
    letterCol = get_column_letter(col)
    fgPreciseSheet.column_dimensions[letterCol].width = colWidth + 5

fgWorkbook.save("FgTesting.xlsx")
